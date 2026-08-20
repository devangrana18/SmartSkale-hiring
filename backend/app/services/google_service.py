import os
import io
import csv
import json
import logging
import urllib.request
import urllib.error
from typing import List, Dict, Any, Optional
import openpyxl

from app.config import settings

logger = logging.getLogger(__name__)

# Column name normalization mapping
FIELD_SYNONYMS = {
    "full_name": ["name", "full name", "employee name", "candidate name", "intern name", "candidate_name"],
    "email": ["email", "email address", "e-mail", "mail", "email_id"],
    "phone": ["phone", "phone number", "mobile", "contact", "contact number", "mobile number"],
    "address": ["address", "current address", "permanent address", "residing at", "location", "city"],
    "role": ["role", "job role", "designation", "position", "profile", "job title"],
    "department": ["department", "dept", "domain", "division", "team"],
    "joining_date": ["date", "joining date", "start date", "commencement date", "joining_date", "document date"],
    "status": ["status", "employment status", "active status"],
    "duration": ["duration", "internship duration", "period", "tenure"],
    "start_date": ["start date", "from date", "starting date"],
    "end_date": ["end date", "to date", "completion date", "ending date"],
    "stipend": ["stipend", "salary", "compensation", "remuneration"],
    "reference_number": ["reference number", "ref no", "ref number", "reference_no", "ref. no."],
    "employee_id": ["employee id", "employee_id", "intern id", "intern_id", "emp id", "emp_id"]
}

class GoogleIntegrationService:
    def __init__(self):
        self.file_id = settings.GOOGLE_DRIVE_FILE_ID
        self.sheet_name = settings.GOOGLE_SHEET_NAME
        self.service_account_json = settings.GOOGLE_SERVICE_ACCOUNT_JSON

    def _normalize_header(self, header: str) -> str:
        clean = header.strip().lower().replace("_", " ")
        for canonical, synonyms in FIELD_SYNONYMS.items():
            if clean in synonyms:
                return canonical
        return clean.replace(" ", "_")

    def fetch_via_service_account(self) -> Optional[List[Dict[str, Any]]]:
        """Fetches spreadsheet data using official Google Sheets / Drive API."""
        if not self.service_account_json:
            return None
        
        try:
            from google.oauth2 import service_account
            from googleapiclient.discovery import build

            SCOPES = [
                'https://www.googleapis.com/auth/spreadsheets.readonly',
                'https://www.googleapis.com/auth/drive.readonly'
            ]
            
            creds = None
            if os.path.exists(self.service_account_json):
                creds = service_account.Credentials.from_service_account_file(
                    self.service_account_json, scopes=SCOPES
                )
            else:
                info = json.loads(self.service_account_json)
                creds = service_account.Credentials.from_service_account_info(
                    info, scopes=SCOPES
                )

            service = build('sheets', 'v4', credentials=creds)
            range_name = f"{self.sheet_name}!A1:Z" if self.sheet_name else "A1:Z"
            
            sheet = service.spreadsheets()
            result = sheet.values().get(spreadsheetId=self.file_id, range=range_name).execute()
            rows = result.get('values', [])
            
            if not rows:
                return []
            
            headers = [self._normalize_header(h) for h in rows[0]]
            records = []
            for row in rows[1:]:
                if not any(row):
                    continue
                record = {}
                for idx, h in enumerate(headers):
                    val = row[idx] if idx < len(row) else ""
                    record[h] = str(val).strip() if val is not None else ""
                records.append(record)
            
            logger.info(f"Fetched {len(records)} records via Google Service Account.")
            return records
        except Exception as e:
            logger.warning(f"Google Service Account fetch failed or unconfigured: {e}")
            return None

    def fetch_via_public_export(self) -> Optional[List[Dict[str, Any]]]:
        """Fetches spreadsheet via Google Drive public/shared export endpoint."""
        export_url = f"https://docs.google.com/spreadsheets/d/{self.file_id}/export?format=csv"
        try:
            req = urllib.request.Request(
                export_url,
                headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
            )
            with urllib.request.urlopen(req, timeout=12) as response:
                content = response.read().decode('utf-8', errors='replace')
                
                reader = csv.reader(io.StringIO(content))
                rows = list(reader)
                if not rows:
                    return []
                
                headers = [self._normalize_header(h) for h in rows[0]]
                records = []
                for row in rows[1:]:
                    if not any(str(cell).strip() for cell in row):
                        continue
                    record = {}
                    for idx, h in enumerate(headers):
                        val = row[idx] if idx < len(row) else ""
                        record[h] = str(val).strip()
                    records.append(record)
                
                logger.info(f"Fetched {len(records)} records via Google Sheet export.")
                return records
        except Exception as e:
            logger.warning(f"Google Sheet export fetch failed: {e}")
            return None

    def fetch_from_excel_file(self, file_path_or_bytes: Any) -> List[Dict[str, Any]]:
        """Parses an uploaded or local Excel .xlsx file."""
        if isinstance(file_path_or_bytes, bytes):
            wb = openpyxl.load_workbook(io.BytesIO(file_path_or_bytes), data_only=True)
        else:
            wb = openpyxl.load_workbook(file_path_or_bytes, data_only=True)
            
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        
        headers = [self._normalize_header(str(h or "")) for h in rows[0]]
        records = []
        for row in rows[1:]:
            if not any(row):
                continue
            record = {}
            for idx, h in enumerate(headers):
                val = row[idx] if idx < len(row) else ""
                record[h] = str(val).strip() if val is not None else ""
            records.append(record)
        return records

    def fetch_employees_data(self) -> List[Dict[str, Any]]:
        """
        Multi-tier fetch:
        1. Try official Service Account API
        2. Try direct Google Sheet export URL
        3. Fallback to built-in seed demo data if offline / unavailable
        """
        # Tier 1
        data = self.fetch_via_service_account()
        if data is not None and len(data) > 0:
            return data

        # Tier 2
        data = self.fetch_via_public_export()
        if data is not None and len(data) > 0:
            return data

        # Tier 3: Local fallback
        logger.info("Using local fallback employee demo records.")
        return self.get_demo_seed_data()

    def get_demo_seed_data(self) -> List[Dict[str, Any]]:
        return [
            {
                "full_name": "Aarav Sharma",
                "email": "aarav.sharma1@example.com",
                "phone": "+91 98765 43210",
                "address": "100, Sector 10, Noida, Uttar Pradesh, India",
                "role": "Full Stack Gen AI Intern",
                "department": "Technical Development",
                "joining_date": "01-08-2026",
                "duration": "3 Months",
                "status": "Active",
                "reference_number": "SKL-2026-1001"
            },
            {
                "full_name": "Vihaan Verma",
                "email": "vihaan.verma2@example.com",
                "phone": "+91 98765 43211",
                "address": "101, Sector 11, Noida, Uttar Pradesh, India",
                "role": "Full Stack Gen AI Developer",
                "department": "Technical Development",
                "joining_date": "02-08-2026",
                "duration": "6 Months",
                "status": "Active",
                "reference_number": "SKL-2026-1002"
            },
            {
                "full_name": "Aditya Gupta",
                "email": "aditya.gupta3@example.com",
                "phone": "+91 98765 43212",
                "address": "102, Sector 12, Noida, Uttar Pradesh, India",
                "role": "Full Stack Gen AI Intern",
                "department": "Technical Development",
                "joining_date": "03-08-2026",
                "duration": "3 Months",
                "status": "Active",
                "reference_number": "SKL-2026-1003"
            },
            {
                "full_name": "Ananya Mishra",
                "email": "ananya.mishra4@example.com",
                "phone": "+91 98765 43213",
                "address": "205, Indirapuram, Ghaziabad, Uttar Pradesh, India",
                "role": "UI/UX Product Designer",
                "department": "Design & Experience",
                "joining_date": "05-08-2026",
                "duration": "3 Months",
                "status": "Onboarding",
                "reference_number": "SKL-2026-1004"
            },
            {
                "full_name": "Rohan Deshmukh",
                "email": "rohan.deshmukh5@example.com",
                "phone": "+91 98765 43214",
                "address": "304, Hiranandani Estate, Thane, Maharashtra, India",
                "role": "Backend Engineer (FastAPI/Python)",
                "department": "Technical Development",
                "joining_date": "10-08-2026",
                "duration": "6 Months",
                "status": "Pending",
                "reference_number": "SKL-2026-1005"
            }
        ]

google_service = GoogleIntegrationService()
